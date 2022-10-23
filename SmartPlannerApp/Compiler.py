from operator import contains
import openpyxl
import os
import sys

def compileData(InputDict):
    InputList = []
    for key, value in InputDict.items():
        InputList.append(InputDict.get(key).get("Name"))

    filePath = os.path.join(os.path.dirname(__file__), 'ExcelFiles')
    os.chdir(filePath)
    excelFiles = os.listdir('.')
    for i in range(0, len(excelFiles)):
        wb = openpyxl.load_workbook(excelFiles[i])
        sheet = wb.active
        row = 4
        for j, classes in enumerate(InputList[0:4]):
            sheet.cell(column=1, row=row, value=InputDict.get(classes).get("Name"))
            sheet.cell(column=2, row=row, value=int(InputDict.get(classes).get("Credits")))
            row += 1

        row = 4
        for j, classes in enumerate(InputList[4:8]):
            sheet.cell(column=3, row=row, value=InputDict.get(classes).get("Name"))
            sheet.cell(column=4, row=row, value=int(InputDict.get(classes).get("Credits")))
            row += 1

        row = 4
        for j, classes in enumerate(InputList[8:10]):
            sheet.cell(column=5, row=row, value=InputDict.get(classes).get("Name"))
            sheet.cell(column=6, row=row, value=int(InputDict.get(classes).get("Credits")))
            row += 1        
                    
        row = 13
        for j, classes in enumerate(InputList[10:11]):
            sheet.cell(column=1, row=row, value=InputDict.get(classes).get("Name"))
            sheet.cell(column=2, row=row, value=int(InputDict.get(classes).get("Credits")))
            row += 1  

        wb.save(excelFiles[i])
        print(excelFiles[i] + ' completed.')

    sys.exit()