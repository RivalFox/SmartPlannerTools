from operator import contains
import openpyxl
import os
import sys

def compileData(path, fileName):

    # Open and read text file with classes
    fn = os.path.join(path, fileName)
    with open(fn,'r') as f:
        contents = f.readlines()
        #print(contents)
        while True:
              
            # Filepath where excel template for schedule output lives
            filePath = os.path.join(os.path.dirname(__file__), 'ExcelFiles')
            # Goes inside that folder. 
            os.chdir(filePath)
            # Gets the list of excel files inside the folder. 
            excelFiles = os.listdir('.')

            for i in range(0, len(excelFiles)):
                # This is using the openpyxl module to open the Excel file.
                wb = openpyxl.load_workbook(excelFiles[i])
                # Start adding data to sheet
                sheet = wb.active

                # iterates through first 4 classes and adds them to first semester cells
                row = 4
                for j, classes in enumerate(contents[0:4]):
                    sheet.cell(column=1, row=row, value=classes)
                    row += 1

                row = 4
                for j, classes in enumerate(contents[4:8]):
                    sheet.cell(column=3, row=row, value=classes)
                    row += 1

                row = 4
                for j, classes in enumerate(contents[8:10]):
                    sheet.cell(column=5, row=row, value=classes)
                    row += 1        
                    
                row = 13
                for j, classes in enumerate(contents[10:14]):
                    sheet.cell(column=1, row=row, value=classes)
                    row += 1  

                row = 13
                for j, classes in enumerate(contents[14:18]):
                    sheet.cell(column=3, row=row, value=classes)
                    row += 1  

                row = 13
                for j, classes in enumerate(contents[18:20]):
                    sheet.cell(column=5, row=row, value=classes)
                    row += 1  

                row = 22
                for j, classes in enumerate(contents[20:25]):
                    sheet.cell(column=1, row=row, value=classes)
                    row += 1

                row = 22
                for j, classes in enumerate(contents[25:27]):
                    sheet.cell(column=3, row=row, value=classes)
                    row += 1

                wb.save(excelFiles[i])
                print(excelFiles[i] + ' completed.')
            sys.exit()